#!/usr/bin/env python3
"""
Ableton Live Project Analyzer - OPTIMIERT FÜR GESCHWINDIGKEIT
Analysiert Ableton Live-Projekte und extrahiert verwendete VST-Plugins
"""

import os
import zipfile
import xml.etree.ElementTree as ET
import argparse
import json
import gzip
from pathlib import Path
from typing import Dict, List, Set, Optional
import sys
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class AbletonProjectAnalyzer:
    def __init__(self, project_path: str):
        self.project_path = Path(project_path)
        self.projects = []
        self.all_vsts = set()
        self.lock = threading.Lock()
        
    def find_ableton_projects(self) -> List[Path]:
        """Findet alle Ableton Live-Projekte im angegebenen Verzeichnis"""
        projects = []
        for file_path in self.project_path.rglob("*.als"):
            projects.append(file_path)
        return projects
    
    def extract_project_info(self, project_file: Path) -> Optional[Dict]:
        """Extrahiert Informationen aus einer Ableton-Projektdatei - OPTIMIERT"""
        try:
            # Schnelle Header-Erkennung
            with open(project_file, 'rb') as f:
                header = f.read(4)
            
            # ZIP-Datei (neue Ableton-Versionen)
            if header.startswith(b'PK'):
                return self.extract_from_zip_fast(project_file)
            
            # GZIP-Datei (ältere Ableton-Versionen)  
            elif header.startswith(b'\x1f\x8b'):
                return self.extract_from_gzip_fast(project_file)
            
            # Direkte XML-Datei (sehr alte Versionen)
            elif header.startswith(b'<'):
                return self.extract_from_xml_fast(project_file)
            
            return None
            
        except Exception:
            return None
    
    def extract_from_zip_fast(self, project_file: Path) -> Optional[Dict]:
        """Schnelle ZIP-Extraktion"""
        try:
            with zipfile.ZipFile(project_file, 'r') as zip_file:
                if 'Project.xml' not in zip_file.namelist():
                    return None
                xml_content = zip_file.read('Project.xml')
                root = ET.fromstring(xml_content)
                return self.parse_xml_fast(root, project_file)
        except Exception:
            return None
    
    def extract_from_gzip_fast(self, project_file: Path) -> Optional[Dict]:
        """Schnelle GZIP-Extraktion"""
        try:
            with gzip.open(project_file, 'rb') as f:
                xml_content = f.read()
            root = ET.fromstring(xml_content)
            return self.parse_xml_fast(root, project_file)
        except Exception:
            return None
    
    def extract_from_xml_fast(self, project_file: Path) -> Optional[Dict]:
        """Schnelle XML-Extraktion"""
        try:
            with open(project_file, 'rb') as f:
                xml_content = f.read()
            root = ET.fromstring(xml_content)
            return self.parse_xml_fast(root, project_file)
        except Exception:
            return None
    
    def parse_xml_fast(self, root: ET.Element, project_file: Path) -> Dict:
        """Schnelle XML-Parsing mit Track-Details"""
        project_info = {
            'name': project_file.stem,
            'path': str(project_file),
            'vsts': [],
            'tracks': [],
            'scenes': 0
        }
        
        # VST-Plugins extrahieren
        project_info['vsts'] = self.extract_vsts_fast(root)
        
        # Track-Details extrahieren
        project_info['tracks'] = self.extract_tracks_with_vsts(root)
        
        # Scene Anzahl extrahieren
        project_info['scenes'] = len(root.findall('.//Scene'))
        
        return project_info
    
    def extract_vsts_fast(self, root: ET.Element) -> List[Dict]:
        """Extrahiert VST-Plugin-Informationen - OPTIMIERT"""
        vsts = []
        
        # Suche nach VstPluginInfo-Elementen
        for vst_info in root.findall('.//VstPluginInfo'):
            plugin_data = {}
            
            # Plugin-Name
            plug_name_elem = vst_info.find('PlugName')
            if plug_name_elem is not None and 'Value' in plug_name_elem.attrib:
                plugin_data['name'] = plug_name_elem.attrib['Value']
            else:
                continue  # Überspringe wenn kein Name gefunden
            
            # Dateiname
            file_name_elem = vst_info.find('FileName')
            plugin_data['filename'] = file_name_elem.attrib.get('Value', '') if file_name_elem is not None else ''
            
            # Version
            version_elem = vst_info.find('VstVersion')
            plugin_data['version'] = version_elem.attrib.get('Value', '') if version_elem is not None else ''
            
            # Hersteller
            manufacturer_elem = vst_info.find('Manufacturer')
            if manufacturer_elem is not None and 'Value' in manufacturer_elem.attrib:
                plugin_data['manufacturer'] = manufacturer_elem.attrib['Value']
            else:
                # Fallback: Hersteller aus Dateiname ableiten
                filename = plugin_data['filename']
                if 'Maschine' in filename:
                    plugin_data['manufacturer'] = 'Native Instruments'
                elif 'Serum' in filename:
                    plugin_data['manufacturer'] = 'Xfer Records'
                elif 'Massive' in filename:
                    plugin_data['manufacturer'] = 'Native Instruments'
                elif 'FabFilter' in filename:
                    plugin_data['manufacturer'] = 'FabFilter'
                elif 'iZotope' in filename:
                    plugin_data['manufacturer'] = 'iZotope'
                elif 'Youlean' in filename:
                    plugin_data['manufacturer'] = 'Youlean'
                elif 'AR TG' in filename:
                    plugin_data['manufacturer'] = 'AR TG'
                else:
                    plugin_data['manufacturer'] = 'Unbekannt'
            
            vsts.append(plugin_data)
            self.all_vsts.add(f"{plugin_data['manufacturer']} - {plugin_data['name']}")
        
        return vsts
    
    def extract_tracks_with_vsts(self, root: ET.Element) -> List[Dict]:
        """Extrahiert Track-Details mit verwendeten VSTs"""
        tracks = []
        
        # Suche nach allen Tracks (verschiedene Ableton-Versionen)
        track_elements = []
        
        # Verschiedene Track-Typen in Ableton
        track_types = [
            './/AudioTrack',
            './/MidiTrack', 
            './/ReturnTrack',
            './/MasterTrack',
            './/Track'  # Fallback für alle Tracks
        ]
        
        for track_type in track_types:
            for track in root.findall(track_type):
                if track not in track_elements:  # Vermeide Duplikate
                    track_elements.append(track)
        
        for track in track_elements:
            track_info = {
                'name': 'Unbekannter Track',
                'type': 'Audio',  # Default
                'vsts': []
            }
            
            # Track-Name extrahieren
            name_elem = track.find('Name')
            if name_elem is not None and 'Value' in name_elem.attrib:
                track_info['name'] = name_elem.attrib['Value']
            
            # Track-Typ bestimmen
            if track.tag == 'AudioTrack' or track.find('.//AudioTrack') is not None:
                track_info['type'] = 'Audio'
            elif track.tag == 'MidiTrack' or track.find('.//MidiTrack') is not None:
                track_info['type'] = 'MIDI'
            elif track.tag == 'ReturnTrack' or track.find('.//ReturnTrack') is not None:
                track_info['type'] = 'Return'
            elif track.tag == 'MasterTrack' or track.find('.//MasterTrack') is not None:
                track_info['type'] = 'Master'
            
            # VST-Devices auf diesem Track finden
            track_vsts = []
            
            # Suche in DeviceChain
            for device_chain in track.findall('.//DeviceChain'):
                for vst_info in device_chain.findall('.//VstPluginInfo'):
                    plugin_data = self.extract_vst_from_element(vst_info)
                    if plugin_data:
                        track_vsts.append(plugin_data)
            
            # Suche direkt im Track
            for vst_info in track.findall('.//VstPluginInfo'):
                plugin_data = self.extract_vst_from_element(vst_info)
                if plugin_data and plugin_data not in track_vsts:
                    track_vsts.append(plugin_data)
            
            track_info['vsts'] = track_vsts
            tracks.append(track_info)
        
        return tracks
    
    def extract_vst_from_element(self, vst_info: ET.Element) -> Optional[Dict]:
        """Extrahiert VST-Daten aus einem VstPluginInfo-Element"""
        plugin_data = {}
        
        # Plugin-Name
        plug_name_elem = vst_info.find('PlugName')
        if plug_name_elem is not None and 'Value' in plug_name_elem.attrib:
            plugin_data['name'] = plug_name_elem.attrib['Value']
        else:
            return None
        
        # Dateiname
        file_name_elem = vst_info.find('FileName')
        plugin_data['filename'] = file_name_elem.attrib.get('Value', '') if file_name_elem is not None else ''
        
        # Version
        version_elem = vst_info.find('VstVersion')
        plugin_data['version'] = version_elem.attrib.get('Value', '') if version_elem is not None else ''
        
        # Hersteller
        manufacturer_elem = vst_info.find('Manufacturer')
        if manufacturer_elem is not None and 'Value' in manufacturer_elem.attrib:
            plugin_data['manufacturer'] = manufacturer_elem.attrib['Value']
        else:
            # Fallback: Hersteller aus Dateiname ableiten
            filename = plugin_data['filename']
            if 'Maschine' in filename:
                plugin_data['manufacturer'] = 'Native Instruments'
            elif 'Serum' in filename:
                plugin_data['manufacturer'] = 'Xfer Records'
            elif 'Massive' in filename:
                plugin_data['manufacturer'] = 'Native Instruments'
            elif 'FabFilter' in filename:
                plugin_data['manufacturer'] = 'FabFilter'
            elif 'iZotope' in filename:
                plugin_data['manufacturer'] = 'iZotope'
            elif 'Youlean' in filename:
                plugin_data['manufacturer'] = 'Youlean'
            elif 'AR TG' in filename:
                plugin_data['manufacturer'] = 'AR TG'
            else:
                plugin_data['manufacturer'] = 'Unbekannt'
        
        return plugin_data
    
    def analyze_projects(self, quiet: bool = False, max_workers: int = 16) -> None:
        """Analysiert alle gefundenen Projekte parallel - OPTIMIERT für Geschwindigkeit"""
        print(f"Suche nach Ableton-Projekten in: {self.project_path}")
        project_files = self.find_ableton_projects()
        
        if not project_files:
            print("Keine Ableton-Projekte gefunden!")
            return
        
        print(f"Gefunden: {len(project_files)} Projekt(e)")
        print(f"Starte parallele Analyse mit {max_workers} Threads...")
        
        # Batch-Verarbeitung für bessere Performance
        batch_size = max(1, len(project_files) // max_workers)
        completed = 0
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Batch-Verarbeitung statt einzelne Projekte
            futures = []
            for i in range(0, len(project_files), batch_size):
                batch = project_files[i:i + batch_size]
                future = executor.submit(self.process_batch, batch)
                futures.append(future)
            
            # Sammle Ergebnisse
            for future in as_completed(futures):
                try:
                    batch_results = future.result()
                    completed += len(batch_results)
                    
                    if not quiet:
                        print(f"Fortschritt: {completed}/{len(project_files)} Projekte analysiert")
                    elif completed % 200 == 0:  # Weniger häufige Updates
                        print(f"Fortschritt: {completed}/{len(project_files)} Projekte analysiert...")
                        
                except Exception as e:
                    if not quiet:
                        print(f"Batch-Fehler: {e}")
        
        print(f"Analyse abgeschlossen: {len(self.projects)} Projekte erfolgreich verarbeitet")
    
    def process_batch(self, project_batch: List[Path]) -> List[Dict]:
        """Verarbeitet einen Batch von Projekten"""
        batch_results = []
        for project_file in project_batch:
            try:
                project_info = self.extract_project_info(project_file)
                if project_info:
                    batch_results.append(project_info)
            except Exception:
                # Stille Fehlerbehandlung für bessere Performance
                pass
        
        # Thread-safe Hinzufügung zur Hauptliste
        with self.lock:
            self.projects.extend(batch_results)
        
        return batch_results
    
    def print_summary(self) -> None:
        """Druckt eine Zusammenfassung der Analyse"""
        if not self.projects:
            print("Keine Projekte analysiert!")
            return
        
        total_vsts = sum(len(project['vsts']) for project in self.projects)
        
        print("\n=== ZUSAMMENFASSUNG ===")
        print(f"Analysierte Projekte: {len(self.projects)}")
        print(f"Gesamtanzahl VSTs: {len(self.all_vsts)}")
        
        print("\n=== ALLE VERWENDETEN VSTs ===")
        for vst in sorted(self.all_vsts):
            print(f"- {vst}")
    
    def export_to_json(self, filename: str) -> None:
        """Exportiert die Analyseergebnisse als JSON"""
        data = {
            'timestamp': datetime.now().isoformat(),
            'project_path': str(self.project_path),
            'total_projects': len(self.projects),
            'total_vsts': len(self.all_vsts),
            'projects': self.projects,
            'all_vsts': sorted(list(self.all_vsts))
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"Ergebnisse wurden in {filename} gespeichert!")
    
    def export_vst_lists_recursive(self, base_output_dir: str = "vst_lists") -> None:
        """Exportiert VST-Listen rekursiv für alle Hauptverzeichnisse"""
        base_path = Path(base_output_dir)
        base_path.mkdir(exist_ok=True)
        
        # Gruppiere Projekte nach Hauptverzeichnis (erste Ebene)
        projects_by_main_dir = {}
        for project in self.projects:
            project_path = Path(project['path'])
            # Finde das Hauptverzeichnis (erste Ebene unter dem Suchpfad)
            try:
                relative_path = project_path.relative_to(self.project_path)
                # Nur die erste Ebene verwenden, nicht den kompletten Pfad
                if len(relative_path.parts) > 0:
                    main_dir = relative_path.parts[0]
                else:
                    main_dir = "Root"
            except ValueError:
                # Falls der Pfad nicht relativ ist, verwende den Ordnernamen
                main_dir = project_path.parent.name if project_path.parent.name else "Root"
            
            if main_dir not in projects_by_main_dir:
                projects_by_main_dir[main_dir] = []
            projects_by_main_dir[main_dir].append(project)
        
        total_exported = 0
        
        # Erstelle für jedes Hauptverzeichnis einen Unterordner
        for main_dir, projects in projects_by_main_dir.items():
            main_dir_path = base_path / main_dir
            main_dir_path.mkdir(exist_ok=True)
            
            print(f"\nVerarbeite Hauptverzeichnis: {main_dir} ({len(projects)} Projekte)")
            
            exported_count = 0
            # Erstelle alle VST-Listen für dieses Hauptverzeichnis parallel
            for project in projects:
                # Erstelle Dateiname basierend auf Projektname
                safe_name = "".join(c for c in project['name'] if c.isalnum() or c in (' ', '-', '_')).rstrip()
                txt_filename = f"{safe_name}_VSTs.txt"
                txt_filepath = main_dir_path / txt_filename
                
                # Erstelle VST-Liste mit Track-Details
                vst_list = []
                vst_list.append(f"PROJEKTPFAD: {project['path']}")
                vst_list.append("=" * 60)
                vst_list.append(f"VST-Liste für Projekt: {project['name']}")
                vst_list.append(f"Hauptverzeichnis: {main_dir}")
                vst_list.append(f"Tracks: {len(project['tracks'])}")
                vst_list.append(f"Scenes: {project['scenes']}")
                vst_list.append(f"Anzahl VSTs: {len(project['vsts'])}")
                vst_list.append("=" * 60)
                
                # Track-Details mit VSTs
                if project['tracks']:
                    vst_list.append("TRACK-DETAILS:")
                    vst_list.append("-" * 40)
                    for i, track in enumerate(project['tracks'], 1):
                        vst_list.append(f"{i}. [{track['type']}] {track['name']}")
                        if track['vsts']:
                            for j, vst in enumerate(track['vsts'], 1):
                                vst_info = f"   {j}. {vst['manufacturer']} - {vst['name']}"
                                if vst.get('filename'):
                                    vst_info += f" ({vst['filename']})"
                                if vst.get('version'):
                                    vst_info += f" [Version: {vst['version']}]"
                                vst_list.append(vst_info)
                        else:
                            vst_list.append("   (Keine VSTs auf diesem Track)")
                        vst_list.append("")
                else:
                    vst_list.append("Keine Tracks gefunden.")
                
                vst_list.append("=" * 60)
                vst_list.append("ALLE VSTs (ÜBERSICHT):")
                vst_list.append("-" * 40)
                if project['vsts']:
                    for i, vst in enumerate(project['vsts'], 1):
                        vst_info = f"{i}. {vst['manufacturer']} - {vst['name']}"
                        if vst.get('filename'):
                            vst_info += f" ({vst['filename']})"
                        if vst.get('version'):
                            vst_info += f" [Version: {vst['version']}]"
                        vst_list.append(vst_info)
                else:
                    vst_list.append("Keine VSTs gefunden.")
                
                vst_list.append("=" * 60)
                vst_list.append(f"Erstellt am: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                
                # Schreibe Datei
                try:
                    with open(txt_filepath, 'w', encoding='utf-8') as f:
                        f.write('\n'.join(vst_list))
                    exported_count += 1
                    print(f"  VST-Liste erstellt: {txt_filename}")
                except Exception as e:
                    print(f"  Fehler beim Erstellen von {txt_filename}: {e}")
            
            print(f"  {exported_count} VST-Listen erstellt in: {main_dir}")
            total_exported += exported_count
        
        # Erstelle Zusammenfassung
        self.create_recursive_summary(base_path, projects_by_main_dir)
        
        # Erstelle VST-Bedarf-Liste für neuen PC
        self.create_vst_requirements_list(base_path, projects_by_main_dir)
        
        print(f"\n[OK] Rekursive Inventur abgeschlossen!")
        print(f"Insgesamt {total_exported} VST-Listen erstellt in {len(projects_by_main_dir)} Hauptverzeichnissen")
        print(f"Gespeichert in: {base_path}")
    
    def create_recursive_summary(self, base_path: Path, projects_by_main_dir: dict) -> None:
        """Erstellt eine Zusammenfassung der rekursiven Inventur"""
        summary_file = base_path / "00_INVENTUR_ZUSAMMENFASSUNG.txt"
        
        summary_lines = []
        summary_lines.append("ABLETON STUDIO - VOLLSTÄNDIGE VST-INVENTUR")
        summary_lines.append("=" * 60)
        summary_lines.append(f"Erstellt am: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        summary_lines.append(f"Analysiertes Hauptverzeichnis: {self.project_path}")
        summary_lines.append("=" * 60)
        
        total_projects = 0
        all_vsts_global = set()
        
        for main_dir, projects in projects_by_main_dir.items():
            summary_lines.append(f"\n📁 {main_dir.upper()}")
            summary_lines.append("-" * 40)
            summary_lines.append(f"Projekte: {len(projects)}")
            
            # Sammle alle VSTs für dieses Hauptverzeichnis
            main_dir_vsts = set()
            for project in projects:
                total_projects += 1
                for vst in project['vsts']:
                    vst_key = f"{vst['manufacturer']} - {vst['name']}"
                    main_dir_vsts.add(vst_key)
                    all_vsts_global.add(vst_key)
            
            summary_lines.append(f"Verschiedene VSTs: {len(main_dir_vsts)}")
            if main_dir_vsts:
                summary_lines.append("Verwendete VSTs:")
                for vst in sorted(main_dir_vsts):
                    summary_lines.append(f"  • {vst}")
        
        summary_lines.append("\n" + "=" * 60)
        summary_lines.append("GESAMTÜBERSICHT")
        summary_lines.append("=" * 60)
        summary_lines.append(f"Gesamtprojekte: {total_projects}")
        summary_lines.append(f"Hauptverzeichnisse: {len(projects_by_main_dir)}")
        summary_lines.append(f"Verschiedene VSTs insgesamt: {len(all_vsts_global)}")
        summary_lines.append("\nAlle verwendeten VSTs:")
        for vst in sorted(all_vsts_global):
            summary_lines.append(f"  • {vst}")
        
        summary_lines.append("\n" + "=" * 60)
        summary_lines.append("VERZEICHNISSTRUKTUR")
        summary_lines.append("=" * 60)
        for main_dir, projects in projects_by_main_dir.items():
            summary_lines.append(f"{main_dir}/")
            summary_lines.append(f"  └── {len(projects)} Projektdateien")
        
        # Schreibe Zusammenfassung
        try:
            with open(summary_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(summary_lines))
            print(f"\n[INFO] Zusammenfassung erstellt: {summary_file}")
        except Exception as e:
            print(f"Fehler beim Erstellen der Zusammenfassung: {e}")
    
    def create_vst_requirements_list(self, base_path: Path, projects_by_main_dir: dict) -> None:
        """Erstellt eine VST-Bedarf-Liste für den neuen PC"""
        requirements_file = base_path / "00_VST_BEDARF_FUER_NEUEN_PC.txt"
        
        # Sammle alle VSTs mit Details
        all_vsts_detailed = {}
        total_projects = 0
        
        for main_dir, projects in projects_by_main_dir.items():
            total_projects += len(projects)
            for project in projects:
                for vst in project['vsts']:
                    vst_key = f"{vst['manufacturer']} - {vst['name']}"
                    
                    if vst_key not in all_vsts_detailed:
                        all_vsts_detailed[vst_key] = {
                            'manufacturer': vst['manufacturer'],
                            'name': vst['name'],
                            'filename': vst.get('filename', ''),
                            'version': vst.get('version', ''),
                            'usage_count': 0,
                            'projects': set(),
                            'main_dirs': set()
                        }
                    
                    all_vsts_detailed[vst_key]['usage_count'] += 1
                    all_vsts_detailed[vst_key]['projects'].add(project['name'])
                    all_vsts_detailed[vst_key]['main_dirs'].add(main_dir)
        
        # Sortiere VSTs nach Häufigkeit der Verwendung
        sorted_vsts = sorted(all_vsts_detailed.items(), 
                           key=lambda x: x[1]['usage_count'], reverse=True)
        
        requirements_lines = []
        requirements_lines.append("VST-BEDARF FÜR NEUEN PC")
        requirements_lines.append("=" * 60)
        requirements_lines.append(f"Erstellt am: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        requirements_lines.append(f"Analysierte Projekte: {total_projects}")
        requirements_lines.append(f"Verschiedene VSTs gefunden: {len(all_vsts_detailed)}")
        requirements_lines.append("=" * 60)
        requirements_lines.append("")
        requirements_lines.append("📋 INSTALLATIONSLISTE (sortiert nach Häufigkeit)")
        requirements_lines.append("-" * 60)
        
        for i, (vst_key, details) in enumerate(sorted_vsts, 1):
            requirements_lines.append(f"{i:2d}. {vst_key}")
            requirements_lines.append(f"    Hersteller: {details['manufacturer']}")
            requirements_lines.append(f"    Plugin-Name: {details['name']}")
            if details['filename']:
                requirements_lines.append(f"    Dateiname: {details['filename']}")
            if details['version']:
                requirements_lines.append(f"    Version: {details['version']}")
            requirements_lines.append(f"    Verwendet in: {details['usage_count']} Projekten")
            requirements_lines.append(f"    Hauptverzeichnisse: {', '.join(sorted(details['main_dirs']))}")
            requirements_lines.append("")
        
        requirements_lines.append("=" * 60)
        requirements_lines.append("📊 STATISTIKEN")
        requirements_lines.append("=" * 60)
        
        # Statistiken nach Hersteller
        manufacturer_stats = {}
        for vst_key, details in all_vsts_detailed.items():
            manufacturer = details['manufacturer']
            if manufacturer not in manufacturer_stats:
                manufacturer_stats[manufacturer] = {'count': 0, 'usage': 0}
            manufacturer_stats[manufacturer]['count'] += 1
            manufacturer_stats[manufacturer]['usage'] += details['usage_count']
        
        requirements_lines.append("Nach Hersteller:")
        for manufacturer, stats in sorted(manufacturer_stats.items(), 
                                        key=lambda x: x[1]['usage'], reverse=True):
            requirements_lines.append(f"  • {manufacturer}: {stats['count']} Plugins, {stats['usage']} Verwendungen")
        
        requirements_lines.append("")
        requirements_lines.append("Nach Hauptverzeichnis:")
        for main_dir, projects in projects_by_main_dir.items():
            main_dir_vsts = set()
            for project in projects:
                for vst in project['vsts']:
                    vst_key = f"{vst['manufacturer']} - {vst['name']}"
                    main_dir_vsts.add(vst_key)
            requirements_lines.append(f"  • {main_dir}: {len(main_dir_vsts)} verschiedene VSTs")
        
        requirements_lines.append("")
        requirements_lines.append("=" * 60)
        requirements_lines.append("💡 INSTALLATIONSHINWEISE")
        requirements_lines.append("=" * 60)
        requirements_lines.append("1. Installieren Sie zuerst die häufig verwendeten VSTs")
        requirements_lines.append("2. Prüfen Sie die Kompatibilität mit Ihrer Ableton-Version")
        requirements_lines.append("3. Sichern Sie Ihre Lizenzschlüssel vor der Installation")
        requirements_lines.append("4. Testen Sie die VSTs nach der Installation in Ableton")
        requirements_lines.append("")
        requirements_lines.append("🔗 NÜTZLICHE LINKS:")
        requirements_lines.append("• Native Instruments: https://www.native-instruments.com")
        requirements_lines.append("• FabFilter: https://www.fabfilter.com")
        requirements_lines.append("• iZotope: https://www.izotope.com")
        requirements_lines.append("• Youlean: https://youlean.co")
        
        # Schreibe VST-Bedarf-Liste
        try:
            with open(requirements_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(requirements_lines))
            print(f"\n[INFO] VST-Bedarf-Liste erstellt: {requirements_file}")
        except Exception as e:
            print(f"Fehler beim Erstellen der VST-Bedarf-Liste: {e}")
    
    def export_to_excel(self, filename: str = "ableton_vst_analysis.xlsx") -> None:
        """Exportiert die Analyseergebnisse als übersichtliche Excel-Tabelle"""
        try:
            wb = Workbook()
            
            # Entferne Standard-Sheet
            wb.remove(wb.active)
            
            # 1. Project Overview
            self.create_project_overview_sheet(wb)
            
            # 2. VST Overview
            self.create_vst_overview_sheet(wb)
            
            # 3. Track Details
            self.create_track_details_sheet(wb)
            
            # 4. VST Requirements for New PC
            self.create_vst_requirements_sheet(wb)
            
            # 5. Statistics
            self.create_statistics_sheet(wb)
            
            # Speichere Excel-Datei
            wb.save(filename)
            print(f"\n[INFO] Excel-Analyse erstellt: {filename}")
            
        except Exception as e:
            print(f"Fehler beim Erstellen der Excel-Datei: {e}")
    
    def create_project_overview_sheet(self, wb: Workbook) -> None:
        """Creates Project Overview Sheet"""
        ws = wb.create_sheet("Project Overview")
        
        # Header
        headers = ["Project", "Path", "Tracks", "Scenes", "VSTs", "Main Directory"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Daten
        for row, project in enumerate(self.projects, 2):
            ws.cell(row=row, column=1, value=project['name'])
            ws.cell(row=row, column=2, value=project['path'])
            ws.cell(row=row, column=3, value=len(project['tracks']))
            ws.cell(row=row, column=4, value=project['scenes'])
            ws.cell(row=row, column=5, value=len(project['vsts']))
            
            # Hauptverzeichnis extrahieren
            try:
                project_path = Path(project['path'])
                relative_path = project_path.relative_to(self.project_path)
                main_dir = relative_path.parts[0] if relative_path.parts else "Root"
            except ValueError:
                main_dir = project_path.parent.name if project_path.parent.name else "Root"
            
            ws.cell(row=row, column=6, value=main_dir)
        
        # Auto-fit Spalten
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_vst_overview_sheet(self, wb: Workbook) -> None:
        """Creates VST Overview Sheet"""
        ws = wb.create_sheet("VST Overview")
        
        # Collect all VSTs with details
        vst_data = []
        for project in self.projects:
            for vst in project['vsts']:
                vst_data.append({
                    'Project': project['name'],
                    'Manufacturer': vst['manufacturer'],
                    'VST Name': vst['name'],
                    'Filename': vst.get('filename', ''),
                    'Version': vst.get('version', '')
                })
        
        # Header
        headers = ["Project", "Manufacturer", "VST Name", "Filename", "Version"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, vst in enumerate(vst_data, 2):
            ws.cell(row=row, column=1, value=vst['Project'])
            ws.cell(row=row, column=2, value=vst['Manufacturer'])
            ws.cell(row=row, column=3, value=vst['VST Name'])
            ws.cell(row=row, column=4, value=vst['Filename'])
            ws.cell(row=row, column=5, value=vst['Version'])
        
        # Auto-fit Spalten
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_track_details_sheet(self, wb: Workbook) -> None:
        """Creates Track Details Sheet"""
        ws = wb.create_sheet("Track Details")
        
        # Collect all track data
        track_data = []
        for project in self.projects:
            for track in project['tracks']:
                track_data.append({
                    'Project': project['name'],
                    'Track Name': track['name'],
                    'Track Type': track['type'],
                    'VST Count': len(track['vsts']),
                    'VSTs': ', '.join([f"{vst['manufacturer']} - {vst['name']}" for vst in track['vsts']])
                })
        
        # Header
        headers = ["Project", "Track Name", "Track Type", "VST Count", "VSTs"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, track in enumerate(track_data, 2):
            ws.cell(row=row, column=1, value=track['Project'])
            ws.cell(row=row, column=2, value=track['Track Name'])
            ws.cell(row=row, column=3, value=track['Track Type'])
            ws.cell(row=row, column=4, value=track['VST Count'])
            ws.cell(row=row, column=5, value=track['VSTs'])
        
        # Auto-fit Spalten
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_vst_requirements_sheet(self, wb: Workbook) -> None:
        """Creates VST Requirements Sheet"""
        ws = wb.create_sheet("VST Requirements for New PC")
        
        # Collect VST statistics
        vst_stats = {}
        for project in self.projects:
            for vst in project['vsts']:
                vst_key = f"{vst['manufacturer']} - {vst['name']}"
                if vst_key not in vst_stats:
                    vst_stats[vst_key] = {
                        'Manufacturer': vst['manufacturer'],
                        'VST Name': vst['name'],
                        'Filename': vst.get('filename', ''),
                        'Version': vst.get('version', ''),
                        'Usage Count': 0,
                        'Projects': set()
                    }
                vst_stats[vst_key]['Usage Count'] += 1
                vst_stats[vst_key]['Projects'].add(project['name'])
        
        # Sort by frequency
        sorted_vsts = sorted(vst_stats.items(), key=lambda x: x[1]['Usage Count'], reverse=True)
        
        # Header
        headers = ["Rank", "Manufacturer", "VST Name", "Filename", "Version", "Usage Count", "Projects"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for rank, (vst_key, data) in enumerate(sorted_vsts, 1):
            row = rank + 1
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=data['Manufacturer'])
            ws.cell(row=row, column=3, value=data['VST Name'])
            ws.cell(row=row, column=4, value=data['Filename'])
            ws.cell(row=row, column=5, value=data['Version'])
            ws.cell(row=row, column=6, value=data['Usage Count'])
            ws.cell(row=row, column=7, value=', '.join(sorted(data['Projects'])))
        
        # Auto-fit Spalten
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_statistics_sheet(self, wb: Workbook) -> None:
        """Creates Statistics Sheet"""
        ws = wb.create_sheet("Statistics")
        
        # Calculate statistics
        total_projects = len(self.projects)
        total_vsts = len(self.all_vsts)
        total_tracks = sum(len(project['tracks']) for project in self.projects)
        
        # Manufacturer statistics
        manufacturer_stats = {}
        for project in self.projects:
            for vst in project['vsts']:
                manufacturer = vst['manufacturer']
                if manufacturer not in manufacturer_stats:
                    manufacturer_stats[manufacturer] = 0
                manufacturer_stats[manufacturer] += 1
        
        # Header
        ws.cell(row=1, column=1, value="ABLETON STUDIO - VST ANALYSIS STATISTICS")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        
        # General statistics
        stats_data = [
            ("Total Projects", total_projects),
            ("Different VSTs", total_vsts),
            ("Total Tracks", total_tracks),
            ("Average VSTs per Project", round(sum(len(p['vsts']) for p in self.projects) / total_projects, 2) if total_projects > 0 else 0),
            ("Average Tracks per Project", round(total_tracks / total_projects, 2) if total_projects > 0 else 0)
        ]
        
        for row, (label, value) in enumerate(stats_data, 3):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Manufacturer statistics
        ws.cell(row=9, column=1, value="VST MANUFACTURER STATISTICS")
        ws.cell(row=9, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=10, column=1, value="Manufacturer")
        ws.cell(row=10, column=2, value="VST Count")
        ws.cell(row=10, column=1).font = Font(bold=True)
        ws.cell(row=10, column=2).font = Font(bold=True)
        
        for row, (manufacturer, count) in enumerate(sorted(manufacturer_stats.items(), key=lambda x: x[1], reverse=True), 11):
            ws.cell(row=row, column=1, value=manufacturer)
            ws.cell(row=row, column=2, value=count)
        
        # Auto-fit Spalten
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

def main():
    parser = argparse.ArgumentParser(description='Ableton Live Projekt-Analyzer - OPTIMIERT')
    parser.add_argument('path', help='Pfad zu den Ableton-Projekten')
    parser.add_argument('--json', help='Exportiere Ergebnisse als JSON')
    parser.add_argument('--txt', action='store_true', help='Exportiere VST-Listen als Textdateien')
    parser.add_argument('--excel', help='Exportiere Ergebnisse als Excel-Datei')
    parser.add_argument('--recursive', action='store_true', help='Rekursive Analyse mit Unterverzeichnissen')
    parser.add_argument('--quiet', action='store_true', help='Reduzierte Ausgabe')
    parser.add_argument('--workers', type=int, default=16, help='Anzahl paralleler Threads (Standard: 16)')
    
    args = parser.parse_args()
    
    analyzer = AbletonProjectAnalyzer(args.path)
    
    analyzer.analyze_projects(quiet=args.quiet, max_workers=args.workers)
    analyzer.print_summary()
    
    if args.json:
        analyzer.export_to_json(args.json)
    
    if args.txt:
        if args.recursive:
            analyzer.export_vst_lists_recursive("vst_lists")
        else:
            print("Verwenden Sie --recursive für VST-Listen-Export")
    
    if args.excel:
        analyzer.export_to_excel(args.excel)

if __name__ == "__main__":
    main()